"""
fill_excel_ja.py
读取 Excel 文件，将 B 列中文翻译为日语，填入 E 列（日）。
用法：python fill_excel_ja.py <excel_path> [--style formal|marketing]
"""
import sys
import os
from pathlib import Path

# 确保能 import orchestrator
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from dotenv import load_dotenv

load_dotenv()


def fill_japanese(excel_path: str, style: str = "formal"):
    import orchestrator

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 找到列索引
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    zh_col = None
    ja_col = None
    for idx, val in enumerate(header_row, 1):
        if val == "中":
            zh_col = idx
        elif val == "日":
            ja_col = idx

    if zh_col is None or ja_col is None:
        print(f'[ERROR] 找不到"中"或"日"列，表头为: {header_row}')
        return

    print(f"中文列: {zh_col}，日语列: {ja_col}")
    print()

    filled = 0
    for row_num in range(2, ws.max_row + 1):
        cell_zh = ws.cell(row=row_num, column=zh_col)
        cell_ja = ws.cell(row=row_num, column=ja_col)

        zh_text = cell_zh.value
        if not zh_text or not str(zh_text).strip():
            continue

        zh_text = str(zh_text).strip()
        print(f"  行{row_num} 翻译: {zh_text[:40]} ...", end=" ", flush=True)

        result = orchestrator.translate(zh_text, style=style)
        cell_ja.value = result.final
        filled += 1
        print(f"→ {result.final[:40]}  [评分:{result.score}]")

    wb.save(excel_path)
    print(f"\n完成：共填入 {filled} 条日语译文")
    print(f"已保存：{excel_path}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    style = "formal"
    if "--style" in sys.argv:
        idx = sys.argv.index("--style")
        style = sys.argv[idx + 1]

    if not path:
        print("用法: python fill_excel_ja.py <excel_path> [--style formal|marketing]")
        sys.exit(1)

    fill_japanese(path, style=style)
